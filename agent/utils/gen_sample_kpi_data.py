"""One-shot generator for the 5 sample input files the KPI agent needs.

Run: `python -m utils.gen_sample_kpi_data` from the agent/ dir.
Writes into ../Internal Document (Dummy)/. Values are chosen to reproduce
the headline-KPI table (L01..R02) used in the project brief.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from config import PROJECT_ROOT

DOCS = PROJECT_ROOT / "Internal Document (Dummy)"


def _write(filename: str, sheet: str, rows: list[list]) -> None:
    path = DOCS / filename
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    for r in rows:
        ws.append(r)
    for i, _ in enumerate(rows[0], start=1):
        ws.column_dimensions[get_column_letter(i)].width = 20
    wb.save(path)
    print(f"wrote {path}")


def bank_balances() -> None:
    # Targets: L01=1,248.50  L02=417.30  L03=HDFC 52.4%
    _write(
        "Bank_Balances.xlsx",
        "Balances",
        [
            ["bank",        "currency", "balance_inr_mn", "as_of_date"],
            ["HDFC Bank",   "INR",      280.00,           "2026-04-14"],
            ["HDFC Bank",   "USD",      374.00,           "2026-04-14"],
            ["ICICI Bank",  "INR",      137.30,           "2026-04-14"],
            ["ICICI Bank",  "USD",      150.00,           "2026-04-14"],
            ["SBI",         "USD",      200.00,           "2026-04-14"],
            ["Citibank",    "USD",      107.20,           "2026-04-14"],
        ],
    )


def debt_schedule() -> None:
    # Targets: D01 = 2556 / 1800 = 1.42x   D02 = 922 / 1400 = 65.86%
    _write(
        "Debt_Schedule.xlsx",
        "Debt",
        [
            ["facility",      "type",      "sanctioned_inr_mn", "outstanding_inr_mn", "ebitda_ttm_inr_mn", "covenant"],
            ["ECB Tranche 1", "Term Loan", 2500,                 1800,                 1800,                 "Net Debt/EBITDA ≤ 2.50x"],
            ["ECB Tranche 2", "Term Loan",  900,                  756,                    0,                 ""],
            ["CC Facility",   "Fund",       800,                  527,                    0,                 ""],
            ["LC/BG Limit",   "Non-Fund",   600,                  395,                    0,                 ""],
        ],
    )


def compliance_log() -> None:
    # Target: C01 = 14/15 = 93.33% (AMBER), 1 ECB-2 filing late Aug-25 by 3 days
    rows = [["filing_type", "due_date", "filed_date", "days_late", "penalty"]]
    months = [
        ("2025-04-07", "2025-04-05"), ("2025-05-07", "2025-05-06"),
        ("2025-06-07", "2025-06-06"), ("2025-07-07", "2025-07-05"),
        ("2025-08-07", "2025-08-10"),  # ← late by 3 days
        ("2025-09-07", "2025-09-05"), ("2025-10-07", "2025-10-06"),
        ("2025-11-07", "2025-11-06"), ("2025-12-07", "2025-12-05"),
        ("2026-01-07", "2026-01-05"), ("2026-02-07", "2026-02-06"),
        ("2026-03-07", "2026-03-05"),
    ]
    for due, filed in months:
        days_late = 3 if due == "2025-08-07" else 0
        rows.append(["ECB-2 Return", due, filed, days_late, "None"])
    rows.append(["FLA Return",      "2025-07-15", "2025-07-14", 0, "None"])
    rows.append(["APR",             "2025-09-30", "2025-09-28", 0, "None"])
    rows.append(["FC-GPR",          "2025-06-30", "2025-06-28", 0, "None"])
    _write("Compliance_Log.xlsx", "Log", rows)


def hedge_effectiveness() -> None:
    # Target: C03 = 100%, 11 CFH designations all 80-125%
    rows = [["designation_id", "test_date", "dollar_offset_pct", "pass_fail"]]
    data = [
        ("CFH-USD-2025-01",  98.4), ("CFH-USD-2025-02", 102.7),
        ("CFH-USD-2025-03",  96.1), ("CFH-USD-2025-04", 108.3),
        ("CFH-USD-2025-05",  92.5), ("CFH-USD-2025-06", 104.8),
        ("CFH-USD-2025-07", 110.2), ("CFH-USD-2025-08",  88.9),
        ("CFH-EUR-2025-01", 101.4), ("CFH-EUR-2025-02",  97.6),
        ("CFH-EUR-2025-03", 105.5),
    ]
    for did, offset in data:
        rows.append([did, "2026-03-31", offset, "PASS"])
    _write("Hedge_Effectiveness.xlsx", "CFH_Tests", rows)


def fx_budget() -> None:
    # Target: F05 = 4.2 / 3.5 = 1.20x (current quarter = Q4 FY26)
    _write(
        "FX_Budget.xlsx",
        "Budget",
        [
            ["quarter",   "budget_inr_cr", "realised_inr_cr", "current"],
            ["Q1 FY26",    3.2,              3.4,              "N"],
            ["Q2 FY26",    3.3,              3.1,              "N"],
            ["Q3 FY26",    3.4,              3.8,              "N"],
            ["Q4 FY26",    3.5,              4.2,              "Y"],
        ],
    )


if __name__ == "__main__":
    DOCS.mkdir(exist_ok=True)
    bank_balances()
    debt_schedule()
    compliance_log()
    hedge_effectiveness()
    fx_budget()
    print("done.")
